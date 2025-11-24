import os
import sqlite3
import hashlib
import logging
import random
from datetime import datetime
from typing import Tuple, Dict

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import piexif
    PIEXIF_AVAILABLE = True
except ImportError:
    PIEXIF_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core.file_system import AndroidFileSystem
from utils.binary_utils import set_file_timestamp

class MediaEngine:
    def __init__(self, fs: AndroidFileSystem, logger: logging.Logger):
        self.fs = fs
        self.logger = logger
        if not PIL_AVAILABLE:
            self.logger.warning("Pillow (PIL) not found. Image generation will be skipped.")

    def _to_deg(self, value, loc):
        if value < 0: loc_value = loc[1]
        else: loc_value = loc[0]
        abs_value = abs(value)
        deg = int(abs_value)
        t1 = (abs_value - deg) * 60
        min = int(t1)
        sec = round((t1 - min) * 60 * 10000)
        return (deg, 1), (min, 1), (sec, 10000), loc_value

    def generate_image_file(self, filename: str, timestamp: datetime, location: Tuple[float, float] = None):
        if not PIL_AVAILABLE: return
        main_path = self.fs.get_path("dcim") / filename
        thumb_path = self.fs.get_path("thumbnails") / filename
        try:
            # Improvement #2: Generate Random Noise instead of solid color
            # This looks more like real data in a hex editor/preview
            img = Image.effect_mandelbrot((400, 300), (0, 0, 400, 300), 100)
            # Fallback if effect not available or purely random noise desired:
            if random.random() < 0.5:
                # Create random pixel data
                pixels = bytes([random.randint(0, 255) for _ in range(400 * 300 * 3)])
                img = Image.frombytes('RGB', (400, 300), pixels)
            
            draw = ImageDraw.Draw(img)
            
            # Visual Text Overlay
            text = f"IMG: {filename}\nDate: {timestamp}"
            if location: text += f"\nLat: {location[0]:.4f}\nLon: {location[1]:.4f}"
            
            try:
                # Add a semi-transparent box for text legibility
                draw.rectangle([10, 10, 200, 80], fill=(0, 0, 0))
                draw.text((15, 15), text, fill=(255, 255, 255))
            except Exception: pass
            
            # Save initial image
            img.save(main_path, "JPEG", quality=85)
            
            # Real EXIF Injection
            if PIEXIF_AVAILABLE and location:
                exif_dict = {"GPS": {}}
                lat_deg = self._to_deg(location[0], ["N", "S"])
                lon_deg = self._to_deg(location[1], ["E", "W"])
                
                exif_dict["GPS"][piexif.GPSIFD.GPSLatitude] = [lat_deg[0], lat_deg[1], lat_deg[2]]
                exif_dict["GPS"][piexif.GPSIFD.GPSLatitudeRef] = lat_deg[3]
                exif_dict["GPS"][piexif.GPSIFD.GPSLongitude] = [lon_deg[0], lon_deg[1], lon_deg[2]]
                exif_dict["GPS"][piexif.GPSIFD.GPSLongitudeRef] = lon_deg[3]
                
                exif_bytes = piexif.dump(exif_dict)
                img.save(main_path, "JPEG", exif=exif_bytes)

            set_file_timestamp(main_path, timestamp)
            
            # Generate Thumbnail
            img.resize((320, 240)).save(thumb_path, "JPEG")
            set_file_timestamp(thumb_path, timestamp)
            
        except Exception as e: self.logger.error(f"Error generating image {filename}: {e}")

    def build_media_store_db(self):
        dcim = self.fs.get_path("dcim")
        db_path = self.fs.get_path("media_db")
        db_path.mkdir(parents=True, exist_ok=True)
        try:
            with sqlite3.connect(db_path / "external.db") as conn:
                c = conn.cursor()
                c.execute("CREATE TABLE IF NOT EXISTS files (_id INTEGER PRIMARY KEY, _data TEXT, date_added INTEGER, media_type INTEGER, mime_type TEXT)")
                if dcim.exists():
                    for f in dcim.iterdir():
                        if f.suffix.lower() in ['.jpg', '.jpeg', '.png']:
                            ts = int(f.stat().st_mtime)
                            c.execute("INSERT INTO files (_data, date_added, media_type, mime_type) VALUES (?, ?, ?, ?)", (str(f), ts, 1, "image/jpeg"))
                conn.commit()
        except sqlite3.Error: pass

    def generate_financial_receipts(self, installed_apps: Dict[str, str], timestamp: datetime):
        path = self.fs.get_path("downloads")
        path.mkdir(parents=True, exist_ok=True)
        finance_apps = ["PayPal", "Cash App", "Venmo", "Coinbase", "Amazon"]
        
        # Determine if we should generate a receipt today
        if random.random() < 0.2:
            for fin in finance_apps:
                # Check if app is installed (loose matching)
                is_installed = any(fin.lower() in app_key.lower() for app_key in installed_apps.keys())
                
                if is_installed:
                    filename = f"{fin}_Receipt_{random.randint(10000,99999)}.pdf"
                    file_path = path / filename
                    try:
                        with open(file_path, "wb") as f:
                            f.write(b"%PDF-1.5\n")
                            f.write(f"Transaction confirmed for {fin}\nAmount: ${random.uniform(50, 500):.2f}\nDate: {timestamp}".encode())
                            f.write(b"\n%%EOF")
                        
                        # Improvement #7: Correct timestamping
                        set_file_timestamp(file_path, timestamp)
                        return # One receipt per trigger is enough
                    except OSError: pass

    def generate_download_manager_db(self):
        dl_path = self.fs.get_path("downloads")
        db_path = self.fs.get_path("data") / "com.android.providers.downloads" / "databases"
        db_path.mkdir(parents=True, exist_ok=True)
        try:
            with sqlite3.connect(db_path / "downloads.db") as conn:
                c = conn.cursor()
                c.execute("CREATE TABLE IF NOT EXISTS downloads (_id INTEGER PRIMARY KEY, uri TEXT, _data TEXT, mimetype TEXT, title TEXT, description TEXT)")
                if dl_path.exists():
                    for f in dl_path.iterdir():
                        if f.is_file():
                            uri = f"https://mail.google.com/mail/u/0?ui=2&ik=c12345&view=att&th=123&attid=0.1&disp=safe&zw&name={f.name}"
                            c.execute("INSERT INTO downloads (uri, _data, title, mimetype) VALUES (?, ?, ?, ?)",
                                      (uri, str(f), f.name, "application/octet-stream"))
                conn.commit()
        except sqlite3.Error: pass

    def generate_thumbnail_cache(self):
        path = self.fs.get_path("thumbnails")
        path.mkdir(parents=True, exist_ok=True)
        for i in [3, 4]:
            filename = f".thumbdata3--{random.randint(1000000000, 9999999999)}"
            try:
                with open(path / filename, "wb") as f:
                    f.write(b"\x01\x00\x00\x00")
                    f.write(os.urandom(1024 * 1024))
            except OSError: pass

    def generate_office_docs(self):
        doc_path = self.fs.get_path("sdcard") / "Documents"
        doc_path.mkdir(parents=True, exist_ok=True)
        if OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financials"
            ws['A1'] = "Account"; ws['B1'] = "Balance"
            ws['A2'] = "Offshore"; ws['B2'] = 50000
            try: wb.save(doc_path / "Q3_Financials.xlsx")
            except: pass
        try:
            with open(doc_path / "Meeting_Notes.docx", "wb") as f:
                f.write(b"PK\x03\x04") 
                f.write(b"fake_word_content_xml_structure_here")
        except OSError: pass